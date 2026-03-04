"""
Pass 2: Fix Safety & Security and other resources where the 'url' field contains
template names (e.g. "T1 - Actor Mapping...") instead of actual URLs.
Also clean up N/A, empty, and placeholder urls.
"""
import openpyxl
import json
import os
import re

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ROOT = os.path.dirname(BASE)
XLSM = os.path.join(ROOT, "source-spreadsheet.xlsm")
DATA_JSON = os.path.join(BASE, "src", "data", "process-data.json")

# Extract hyperlinks
wb = openpyxl.load_workbook(XLSM, data_only=False)
link_map = {}
for sname in wb.sheetnames:
    ws = wb[sname]
    for row in ws.iter_rows():
        for cell in row:
            if cell.hyperlink and cell.hyperlink.target:
                target = cell.hyperlink.target
                if not target or target.startswith("#"):
                    continue
                val = str(cell.value).strip() if cell.value else ""
                if val:
                    link_map[val] = target

print(f"Loaded {len(link_map)} hyperlinks")

with open(DATA_JSON, "r", encoding="utf-8") as f:
    data = json.load(f)

fixed = 0
cleaned = 0

def is_real_url(u):
    return u and (u.startswith("http://") or u.startswith("https://") or u.startswith("mailto:"))

def find_best_link(text):
    """Try to find a Box/RescueNet link for template references in text."""
    # Try exact match first
    if text in link_map:
        return link_map[text]
    # Try first line (multi-line template lists)
    first_line = text.split("\n")[0].split("\r")[0].strip()
    if first_line in link_map:
        return link_map[first_line]
    # Try matching "T<number> - ..." pattern
    t_match = re.match(r'(T\d+\w?\s*-\s*[^,\n\r]+)', text)
    if t_match:
        tname = t_match.group(1).strip()
        if tname in link_map:
            return link_map[tname]
        # Try partial
        for key, url in link_map.items():
            if key.startswith(tname[:20]):
                return url
    return None

def fix_resource(res):
    global fixed, cleaned
    name = res.get("name", "").strip()
    url = res.get("url", "").strip()
    
    # Already good
    if is_real_url(url):
        return
    
    # Clean up junk
    if url in ("N/A", "link?", "Resources?", "Do we have global template?", 
               "TBD: step tool kits", "RED", ""):
        res["url"] = ""
        cleaned += 1
        # Still try to find a link by name
        found = find_best_link(name)
        if found and is_real_url(found):
            res["url"] = found
            fixed += 1
        return
    
    # URL field has template names — try to resolve
    found = find_best_link(url)
    if found and is_real_url(found):
        res["url"] = found
        fixed += 1
        return
    
    # Try by resource name
    found = find_best_link(name)
    if found and is_real_url(found):
        res["url"] = found
        fixed += 1
        return
    
    # URL contains a real URL embedded in text
    url_match = re.search(r'(https?://\S+)', url)
    if url_match:
        res["url"] = url_match.group(1)
        fixed += 1
        return
    
    # Give up — clean the url field so UI shows it as plain text
    res["url"] = ""
    cleaned += 1

for sector in data.get("sectors", []):
    for task in sector.get("tasks", []):
        for res in task.get("resources", []):
            fix_resource(res)
        for subtask in task.get("subtasks", []):
            for res in subtask.get("resources", []):
                fix_resource(res)

# Also clean EmU and Preparedness
for svc in data.get("emuServices", []):
    link = svc.get("link", "").strip()
    if link and not is_real_url(link):
        found = find_best_link(link) or find_best_link(svc.get("name", ""))
        if found and is_real_url(found):
            svc["link"] = found
            fixed += 1
        else:
            svc["link"] = ""
            cleaned += 1

for item in data.get("preparednessLibrary", []):
    link = item.get("link", "").strip()
    name = item.get("name", "").strip()
    if link and not is_real_url(link):
        found = find_best_link(link) or find_best_link(name)
        if found and is_real_url(found):
            item["link"] = found
            fixed += 1
        else:
            item["link"] = ""
            cleaned += 1
    elif not link:
        found = find_best_link(name)
        if found and is_real_url(found):
            item["link"] = found
            fixed += 1

with open(DATA_JSON, "w", encoding="utf-8") as f:
    json.dump(data, f, indent=2, ensure_ascii=False)

print(f"Fixed {fixed} more URLs, cleaned {cleaned} junk entries")

# Final count
total_real = 0
total_empty = 0
for sector in data.get("sectors", []):
    for task in sector.get("tasks", []):
        for res in task.get("resources", []):
            if is_real_url(res.get("url", "")):
                total_real += 1
            else:
                total_empty += 1
        for st in task.get("subtasks", []):
            for res in st.get("resources", []):
                if is_real_url(res.get("url", "")):
                    total_real += 1
                else:
                    total_empty += 1

print(f"\nFinal: {total_real} resources with real URLs, {total_empty} without")

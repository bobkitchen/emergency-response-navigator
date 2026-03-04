"""
Extract hyperlinks from source XLSM and inject them into process-data.json.
Matches by resource/service/preparedness name -> hyperlink target.
"""
import openpyxl
import json
import os

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ROOT = os.path.dirname(BASE)
XLSM = os.path.join(ROOT, "source-spreadsheet.xlsm")
DATA_JSON = os.path.join(BASE, "src", "data", "process-data.json")

# Step 1: Extract ALL hyperlinks from the workbook, keyed by cell value
wb = openpyxl.load_workbook(XLSM, data_only=False)
link_map = {}  # name -> url

for sname in wb.sheetnames:
    ws = wb[sname]
    for row in ws.iter_rows():
        for cell in row:
            if cell.hyperlink and cell.hyperlink.target:
                target = cell.hyperlink.target
                if not target or target.startswith("#"):
                    continue
                val = str(cell.value).strip() if cell.value else ""
                if val and val != "None":
                    # Store the full name as key
                    link_map[val] = target
                    # Also store truncated version (for matching)
                    short = val[:80]
                    if short not in link_map:
                        link_map[short] = target

print(f"Extracted {len(link_map)} hyperlinks from XLSM")

# Step 2: Load process-data.json
with open(DATA_JSON, "r", encoding="utf-8") as f:
    data = json.load(f)

# Step 3: Fix resource URLs in sectors->tasks and subtasks
fixed_count = 0
missing = []

def fix_resource(res):
    global fixed_count
    name = res.get("name", "").strip()
    current_url = res.get("url", "").strip()
    
    # If already has a real URL, skip
    if current_url.startswith("http://") or current_url.startswith("https://") or current_url.startswith("mailto:"):
        return
    
    # Try to find in link_map
    if name in link_map:
        res["url"] = link_map[name]
        fixed_count += 1
        return
    
    # Try matching by current url field (sometimes it has the display text)
    if current_url and current_url in link_map:
        res["url"] = link_map[current_url]
        fixed_count += 1
        return
    
    # Try partial match - first 60 chars
    for key, url in link_map.items():
        if name and key.startswith(name[:60]) and len(name) > 5:
            res["url"] = url
            fixed_count += 1
            return
    
    if name and current_url not in ("", "N/A", "link?", "Resources?", "See MEAL tab", 
                                      "See Partnership tab", "See Response Management Plan",
                                      "See Preparedness Library tab", "RED"):
        missing.append(name)

for sector in data.get("sectors", []):
    for task in sector.get("tasks", []):
        for res in task.get("resources", []):
            fix_resource(res)
        for subtask in task.get("subtasks", []):
            for res in subtask.get("resources", []):
                fix_resource(res)

# Step 4: Fix EmU Services links
for svc in data.get("emuServices", []):
    link_text = svc.get("link", "").strip()
    name = svc.get("name", "").strip()
    
    if link_text.startswith("http") or link_text.startswith("mailto"):
        continue
    
    # Try link text as key
    if link_text in link_map:
        svc["link"] = link_map[link_text]
        fixed_count += 1
    elif name in link_map:
        svc["link"] = link_map[name]
        fixed_count += 1

# Step 5: Fix Preparedness Library links
for item in data.get("preparednessLibrary", []):
    name = item.get("name", "").strip()
    link = item.get("link", "").strip()
    
    if link.startswith("http") or link.startswith("mailto"):
        continue
    
    if name in link_map:
        item["link"] = link_map[name]
        fixed_count += 1

# Step 6: Write updated JSON
with open(DATA_JSON, "w", encoding="utf-8") as f:
    json.dump(data, f, indent=2, ensure_ascii=False)

print(f"Fixed {fixed_count} URLs in process-data.json")
if missing:
    print(f"\n{len(missing)} resources still without URLs:")
    for m in sorted(set(missing))[:30]:
        print(f"  - {m[:80]}")
